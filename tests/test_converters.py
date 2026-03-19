"""导出器单元测试 — 使用构造的 DocumentResult，不依赖 PaddleOCR。"""

import sys
import types
from dataclasses import replace
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

from app.models import BlockResult, BlockType, DocumentResult, PageResult
from app.converters.txt_converter import TxtConverter
from app.converters.rtf_converter import RtfConverter
from app.converters.word_converter import WordConverter
from app.converters.html_converter import HtmlConverter
from app.converters.excel_converter import ExcelConverter
from app.converters.pdf_converter import PdfConverter


class _FakePoint:
    def __init__(self, x, y):
        self.x = x
        self.y = y


class _FakeFont:
    def __init__(self, fontname: str, font_specs: dict[str, object]):
        spec = font_specs[fontname]
        if isinstance(spec, Exception):
            raise spec
        self.fontname = fontname
        self.ascender = spec["ascender"]
        self.descender = spec["descender"]
        self._unit_width = spec.get("unit_width")

    def text_length(self, text: str, fontsize: float = 1) -> float:
        if self._unit_width is None:
            raise RuntimeError("missing width")
        return self._unit_width * fontsize


class _FakePdfPage:
    def __init__(self, fail_callback=None):
        self.calls: list[dict[str, object]] = []
        self._fail_callback = fail_callback

    def insert_text(self, point, text, fontname, fontsize, render_mode):
        call = {
            "point": point,
            "text": text,
            "fontname": fontname,
            "fontsize": fontsize,
            "render_mode": render_mode,
        }
        self.calls.append(call)
        if self._fail_callback is not None:
            self._fail_callback(call, len(self.calls))


def _install_fake_fitz(monkeypatch, font_specs, width_lookup=None):
    width_lookup = width_lookup or {}

    class FakeFont:
        def __init__(self, fontname: str):
            fake_font = _FakeFont(fontname, font_specs)
            self.ascender = fake_font.ascender
            self.descender = fake_font.descender
            self._fontname = fontname
            self._fallback_width = fake_font._unit_width

        def text_length(self, text: str, fontsize: float = 1) -> float:
            width = width_lookup.get((self._fontname, text), self._fallback_width)
            if isinstance(width, Exception):
                raise width
            if width is None:
                raise RuntimeError("missing width")
            return width * fontsize

    def get_text_length(text: str, fontname: str, fontsize: float = 1) -> float:
        spec = font_specs[fontname]
        if isinstance(spec, Exception):
            raise spec
        width = width_lookup.get((fontname, text), spec.get("unit_width"))
        if isinstance(width, Exception):
            raise width
        if width is None:
            raise RuntimeError("missing width")
        return width * fontsize

    fake_fitz = types.SimpleNamespace(
        Font=FakeFont,
        Point=_FakePoint,
        get_text_length=get_text_length,
    )
    monkeypatch.setitem(sys.modules, "fitz", fake_fitz)
    return fake_fitz


def _install_fake_fitz_for_convert(monkeypatch):
    class FakeFont:
        def __init__(self, fontname: str):
            self.ascender = 1.0
            self.descender = -0.2

    class FakeRect:
        def __init__(self, x1, y1, x2, y2):
            self.width = x2 - x1
            self.height = y2 - y1

    class FakePage:
        def __init__(self, width: float, height: float):
            self.rect = types.SimpleNamespace(width=width, height=height)

        def insert_image(self, rect, filename):
            return None

        def insert_text(self, point, text, fontname, fontsize, render_mode):
            return None

    class FakeDoc:
        def __init__(self):
            self.pages: list[FakePage] = []
            self.saved_paths: list[str] = []

        def new_page(self, width, height):
            page = FakePage(width, height)
            self.pages.append(page)
            return page

        def save(self, path: str):
            self.saved_paths.append(path)

        def close(self):
            return None

    docs: list[FakeDoc] = []

    def open_doc(*args, **kwargs):
        doc = FakeDoc()
        docs.append(doc)
        return doc

    fake_fitz = types.SimpleNamespace(
        Font=FakeFont,
        Point=_FakePoint,
        Rect=FakeRect,
        open=open_doc,
        get_text_length=lambda text, fontname, fontsize=1: max(len(text), 1) * fontsize,
    )
    monkeypatch.setitem(sys.modules, "fitz", fake_fitz)
    return docs


@pytest.fixture
def sample_result() -> DocumentResult:
    return DocumentResult(
        source_path=Path("/tmp/test.png"),
        page_count=1,
        pages=[
            PageResult(
                page_index=0,
                width=800,
                height=400,
                blocks=[
                    BlockResult(
                        block_type=BlockType.TITLE,
                        bbox=(10.0, 10.0, 200.0, 40.0),
                        text="Test Title",
                        confidence=0.99,
                    ),
                    BlockResult(
                        block_type=BlockType.PARAGRAPH,
                        bbox=(10.0, 50.0, 400.0, 80.0),
                        text="This is a paragraph of text.",
                        confidence=0.95,
                    ),
                    BlockResult(
                        block_type=BlockType.TABLE,
                        bbox=(10.0, 100.0, 400.0, 200.0),
                        text="",
                        table_cells=[
                            ["Name", "Age", "Score"],
                            ["Alice", "25", "98.5"],
                            ["Bob", "30", "87.0"],
                        ],
                    ),
                ],
            )
        ],
        plain_text="Test Title\nThis is a paragraph of text.",
    )


class TestTxtConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.txt"
        TxtConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert "Test Title" in content
        assert "paragraph" in content


class TestRtfConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.rtf"
        RtfConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert r"\rtf1" in content
        assert "Test Title" in content

    def test_fallback_plain_text(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.pdf"),
            page_count=1,
            pages=[],
            plain_text="Fallback line 1\nFallback line 2",
        )
        out = tmp_path / "fallback.rtf"
        RtfConverter().convert(result, out)
        content = out.read_text()
        assert "Fallback line 1" in content
        assert "Fallback line 2" in content


class TestWordConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.docx"
        WordConverter().convert(sample_result, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_fallback_plain_text(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[PageResult(page_index=0, width=100, height=100, blocks=[])],
            plain_text="Fallback text line",
        )
        out = tmp_path / "fallback.docx"
        WordConverter().convert(result, out)
        assert out.exists()


class TestHtmlConverter:
    def test_creates_file(self, sample_result, tmp_path):
        out = tmp_path / "out.html"
        HtmlConverter().convert(sample_result, out)
        assert out.exists()
        content = out.read_text()
        assert "<h1>" in content
        assert "Test Title" in content
        assert "<table>" in content

    def test_escapes_html(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[
                PageResult(
                    page_index=0, width=100, height=100,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(0, 0, 100, 20),
                            text="<script>alert('xss')</script>",
                        )
                    ],
                )
            ],
            plain_text="",
        )
        out = tmp_path / "xss.html"
        HtmlConverter().convert(result, out)
        content = out.read_text()
        assert "<script>" not in content
        assert "&lt;script&gt;" in content


class TestExcelConverter:
    def test_creates_file_with_table(self, sample_result, tmp_path):
        out = tmp_path / "out.xlsx"
        ExcelConverter().convert(sample_result, out)
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "Table_1" in wb.sheetnames
        ws = wb["Table_1"]
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"

    def test_fallback_no_tables(self, tmp_path):
        result = DocumentResult(
            source_path=Path("/tmp/test.png"),
            page_count=1,
            pages=[
                PageResult(
                    page_index=0, width=100, height=100,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(0, 0, 100, 20),
                            text="Just text",
                        )
                    ],
                )
            ],
            plain_text="Just text",
        )
        out = tmp_path / "fallback.xlsx"
        ExcelConverter().convert(result, out)
        assert out.exists()

        from openpyxl import load_workbook
        wb = load_workbook(str(out))
        assert "OCR Text" in wb.sheetnames


class TestPdfConverter:
    def test_overlay_block_preserves_blank_line_layout(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "china-s": {"ascender": 1.0, "descender": -0.2, "unit_width": 4.0},
                "helv": {"ascender": 1.0, "descender": -0.25, "unit_width": 2.0},
                "cour": {"ascender": 1.0, "descender": -0.25, "unit_width": 2.0},
            },
        )
        page = _FakePdfPage()
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(0.0, 0.0, 20.0, 30.0),
            text="第一行\n\nSecond line",
        )

        PdfConverter._overlay_block(
            page,
            block,
            1.0,
            font_cache={},
            fallback_fonts=("helv", "cour"),
        )

        assert len(page.calls) == 2
        assert page.calls[0]["fontname"] == "china-s"
        assert page.calls[0]["fontsize"] == pytest.approx(5.0)
        assert page.calls[0]["point"].y == pytest.approx(5.0)
        assert page.calls[1]["fontname"] == "helv"
        assert page.calls[1]["point"].y == pytest.approx(28.0)

    def test_overlay_block_falls_back_when_font_metrics_invalid(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "china-s": {"ascender": 0.0, "descender": -0.2, "unit_width": 1.0},
                "helv": {"ascender": 1.0, "descender": -0.25, "unit_width": 1.0},
                "cour": {"ascender": 1.0, "descender": -0.25, "unit_width": 1.0},
            },
        )
        page = _FakePdfPage()
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(0.0, 0.0, 100.0, 20.0),
            text="中文",
        )

        PdfConverter._overlay_block(
            page,
            block,
            1.0,
            font_cache={},
            fallback_fonts=("helv", "china-s", "cour"),
        )

        assert len(page.calls) == 1
        assert page.calls[0]["fontname"] == "china-s"
        assert page.calls[0]["fontsize"] == pytest.approx(14.0)
        assert page.calls[0]["point"].y == pytest.approx(14.0)

    def test_overlay_block_retries_with_fallback_font_and_remeasures_width(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "china-s": {"ascender": 1.0, "descender": -0.2, "unit_width": 8.0},
                "helv": {"ascender": 1.0, "descender": -0.2, "unit_width": 4.0},
                "cour": {"ascender": 1.0, "descender": -0.2, "unit_width": 6.0},
            },
            width_lookup={
                ("china-s", "中文"): 8.0,
                ("helv", "中文"): 4.0,
                ("cour", "中文"): 6.0,
            },
        )

        def fail_on_china_s(call, _index):
            if call["fontname"] == "china-s":
                raise RuntimeError("font glyph missing")

        page = _FakePdfPage(fail_callback=fail_on_china_s)
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(0.0, 0.0, 24.0, 10.0),
            text="中文",
        )

        PdfConverter._overlay_block(
            page,
            block,
            1.0,
            font_cache={},
            fallback_fonts=("helv", "china-s", "cour"),
        )

        assert len(page.calls) == 3
        assert page.calls[-1]["fontname"] == "helv"
        assert page.calls[-1]["fontsize"] == pytest.approx(6.0)

    def test_overlay_block_raises_on_non_font_insert_error(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "helv": {"ascender": 1.0, "descender": -0.2, "unit_width": 2.0},
                "cour": {"ascender": 1.0, "descender": -0.2, "unit_width": 2.0},
            },
        )

        def fail_with_page_error(call, _index):
            raise RuntimeError("page closed")

        page = _FakePdfPage(fail_callback=fail_with_page_error)
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(0.0, 0.0, 30.0, 10.0),
            text="ASCII text",
        )

        with pytest.raises(RuntimeError, match="PDF 文本层写入失败"):
            PdfConverter._overlay_block(
                page,
                block,
                1.0,
                font_cache={},
                fallback_fonts=("helv", "cour"),
            )

    def test_overlay_block_raises_when_all_font_fallbacks_fail(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "china-s": {"ascender": 1.0, "descender": -0.2, "unit_width": 4.0},
                "helv": {"ascender": 1.0, "descender": -0.2, "unit_width": 4.0},
                "cour": {"ascender": 1.0, "descender": -0.2, "unit_width": 4.0},
            },
        )

        def fail_with_glyph_error(call, _index):
            raise RuntimeError("glyph missing")

        page = _FakePdfPage(fail_callback=fail_with_glyph_error)
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(0.0, 0.0, 20.0, 10.0),
            text="中文",
        )

        with pytest.raises(RuntimeError, match="PDF 文本层写入失败"):
            PdfConverter._overlay_block(
                page,
                block,
                1.0,
                font_cache={},
                fallback_fonts=("helv", "china-s", "cour"),
            )

    def test_overlay_block_skips_invalid_bbox(self, monkeypatch):
        _install_fake_fitz(
            monkeypatch,
            font_specs={
                "helv": {"ascender": 1.0, "descender": -0.2, "unit_width": 2.0},
            },
        )
        page = _FakePdfPage()
        overlay_error_state = {"count": 0, "samples": []}
        block = BlockResult(
            block_type=BlockType.PARAGRAPH,
            bbox=(10.0, 10.0, 5.0, 20.0),
            text="ASCII text",
        )

        PdfConverter._overlay_block(
            page,
            block,
            1.0,
            font_cache={},
            fallback_fonts=("helv",),
            overlay_error_state=overlay_error_state,
        )

        assert page.calls == []
        assert overlay_error_state == {
            "count": 1,
            "samples": ["invalid-size bbox=(10.0, 10.0, 5.0, 20.0) scale=1.0"],
        }

    def test_convert_raises_summary_for_invalid_blocks(self, monkeypatch, tmp_path):
        docs = _install_fake_fitz_for_convert(monkeypatch)
        result = DocumentResult(
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
            page_count=1,
            pages=[
                PageResult(
                    page_index=0,
                    width=100,
                    height=100,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(20.0, 20.0, 10.0, 30.0),
                            text="invalid block",
                        )
                    ],
                )
            ],
        )

        converter = PdfConverter()
        converter.strict_text_layer = True
        with pytest.raises(RuntimeError, match="PDF 文本层存在非法块： count=1"):
            converter.convert(result, tmp_path / "invalid.pdf")

        assert docs
        assert docs[0].saved_paths == []

    def test_creates_searchable_pdf(self, sample_result, tmp_path):
        pytest.importorskip("fitz")
        result = replace(
            sample_result,
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
        )
        out = tmp_path / "out.pdf"
        PdfConverter().convert(result, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_text_layer_is_extractable(self, sample_result, tmp_path):
        """文字层必须可被 PDF 阅读器提取（可选取/复制）。"""
        fitz = pytest.importorskip("fitz")

        result = replace(
            sample_result,
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
        )
        out = tmp_path / "out.pdf"
        PdfConverter().convert(result, out)

        doc = fitz.open(str(out))
        text = doc[0].get_text()
        doc.close()
        assert "Test Title" in text
        assert "paragraph" in text

    def test_cjk_text_layer(self, tmp_path):
        """中文文字层不能变成问号，必须保留原始 Unicode。"""
        fitz = pytest.importorskip("fitz")

        cjk_result = DocumentResult(
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
            page_count=1,
            pages=[
                PageResult(
                    page_index=0,
                    width=800,
                    height=400,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(10.0, 10.0, 400.0, 50.0),
                            text="测试中文文本层",
                            confidence=0.95,
                        ),
                    ],
                )
            ],
            plain_text="测试中文文本层",
        )
        out = tmp_path / "cjk.pdf"
        PdfConverter().convert(cjk_result, out)

        doc = fitz.open(str(out))
        text = doc[0].get_text()
        doc.close()
        assert "?" not in text, f"CJK text corrupted to question marks: {text!r}"
        assert "测试中文文本层" in text

    def test_multiline_paragraph_text_layer(self, tmp_path):
        """多行段落的每一行都必须出现在文字层中。"""
        fitz = pytest.importorskip("fitz")

        result = DocumentResult(
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
            page_count=1,
            pages=[
                PageResult(
                    page_index=0,
                    width=800,
                    height=600,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(10.0, 10.0, 700.0, 200.0),
                            text="First line of paragraph\nSecond line of paragraph\nThird line here",
                            confidence=0.95,
                        ),
                    ],
                )
            ],
            plain_text="First line\nSecond line\nThird line",
        )
        out = tmp_path / "multiline.pdf"
        PdfConverter().convert(result, out)

        doc = fitz.open(str(out))
        text = doc[0].get_text()
        doc.close()
        assert "First line" in text
        assert "Second line" in text
        assert "Third line" in text

    def test_mixed_cjk_latin_blocks(self, tmp_path):
        """中英混合块：每个块使用正确的字体。"""
        fitz = pytest.importorskip("fitz")

        result = DocumentResult(
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
            page_count=1,
            pages=[
                PageResult(
                    page_index=0,
                    width=800,
                    height=400,
                    blocks=[
                        BlockResult(
                            block_type=BlockType.TITLE,
                            bbox=(10.0, 10.0, 400.0, 40.0),
                            text="English Title",
                            confidence=0.99,
                        ),
                        BlockResult(
                            block_type=BlockType.PARAGRAPH,
                            bbox=(10.0, 50.0, 400.0, 80.0),
                            text="中文段落内容",
                            confidence=0.95,
                        ),
                    ],
                )
            ],
            plain_text="English Title\n中文段落内容",
        )
        out = tmp_path / "mixed.pdf"
        PdfConverter().convert(result, out)

        doc = fitz.open(str(out))
        text = doc[0].get_text()
        doc.close()
        assert "English Title" in text
        assert "中文段落内容" in text
        assert "?" not in text

    def test_many_blocks_all_present(self, tmp_path):
        """大量块都必须出现在文字层中，不能丢失。"""
        fitz = pytest.importorskip("fitz")

        blocks = []
        for i in range(20):
            blocks.append(
                BlockResult(
                    block_type=BlockType.PARAGRAPH,
                    bbox=(10.0, 10.0 + i * 30, 700.0, 35.0 + i * 30),
                    text=f"Block number {i} text content",
                    confidence=0.9,
                )
            )

        result = DocumentResult(
            source_path=Path(__file__).parent / "fixtures" / "test_en.png",
            page_count=1,
            pages=[
                PageResult(page_index=0, width=800, height=700, blocks=blocks)
            ],
            plain_text="\n".join(b.text for b in blocks),
        )
        out = tmp_path / "many_blocks.pdf"
        PdfConverter().convert(result, out)

        doc = fitz.open(str(out))
        text = doc[0].get_text()
        doc.close()
        for i in range(20):
            assert f"Block number {i}" in text, f"Block {i} missing from text layer"
