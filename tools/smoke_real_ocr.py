"""扫描件格式烟测。

覆盖两种输入：
1. 扫描图片
2. 扫描 PDF（由测试图片合成）
"""

from __future__ import annotations

import json
from pathlib import Path
import sys
from tempfile import TemporaryDirectory

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.core.export_router import create_default_router
from app.core.onnx_engine import OnnxOCREngine
from app.core.ocr_worker import OCRWorker
from app.models.enums import OutputFormat
from app.models.job import OCRJob


def _export_all(doc, out_dir: Path) -> dict[str, object]:
    router = create_default_router()
    outputs: dict[str, object] = {}
    for fmt in OutputFormat:
        converter = router.select_converter(fmt)
        output_path = out_dir / f"result_{fmt.value}{converter.file_extension}"
        converter.convert(doc, output_path)

        item: dict[str, object] = {
            "path": str(output_path),
            "size": output_path.stat().st_size,
        }
        if fmt in (OutputFormat.TXT, OutputFormat.HTML, OutputFormat.RTF):
            item["snippet"] = output_path.read_text(encoding="utf-8")[:160]
        elif fmt is OutputFormat.EXCEL:
            item["sheets"] = load_workbook(output_path).sheetnames

        outputs[fmt.value] = item
    return outputs


def _run_image_flow(image_path: Path, out_dir: Path) -> dict[str, object]:
    doc = OnnxOCREngine(lang="en", speed_mode="server").predict(image_path)
    return {
        "input": str(image_path),
        "plain_text": doc.plain_text[:200],
        "page_count": doc.page_count,
        "block_count": len(doc.pages[0].blocks) if doc.pages else 0,
        "outputs": _export_all(doc, out_dir / "image"),
    }


def _make_scanned_pdf(image_path: Path, pdf_path: Path) -> Path:
    import fitz

    image_doc = fitz.open(str(image_path))
    rect = image_doc[0].rect
    image_doc.close()

    pdf = fitz.open()
    page = pdf.new_page(width=rect.width, height=rect.height)
    page.insert_image(page.rect, filename=str(image_path))
    pdf.save(str(pdf_path))
    pdf.close()
    return pdf_path


def _run_worker(job: OCRJob):
    state: dict[str, object] = {}
    worker = OCRWorker(job)
    worker.finished.connect(lambda doc: state.setdefault("doc", doc))
    worker.error.connect(lambda msg: state.setdefault("error", msg))
    worker.run()
    if "error" in state:
        raise RuntimeError(str(state["error"]))
    return state["doc"]


def _run_scanned_pdf_flow(image_path: Path, out_dir: Path) -> dict[str, object]:
    pdf_path = _make_scanned_pdf(image_path, out_dir / "scan.pdf")
    summary: dict[str, object] = {
        "input": str(pdf_path),
        "formats": {},
    }

    for fmt in OutputFormat:
        job = OCRJob(source_path=pdf_path, output_format=fmt, language="en")
        job._adv_params = {
            "speed_mode": "server",
            "parallel_workers": 1,
            "force_ocr": False,
            "page_start": 1,
            "page_end": 1,
            "render_dpi": 400,
            "pipeline": "auto",
        }
        doc = _run_worker(job)
        fmt_dir = out_dir / f"pdf_{fmt.value}"
        fmt_dir.mkdir(parents=True, exist_ok=True)
        item: dict[str, object] = {
            "plain_text": doc.plain_text[:200],
            "page_count": doc.page_count,
            "block_count": len(doc.pages[0].blocks) if doc.pages else 0,
        }
        try:
            outputs = _export_all(doc, fmt_dir)
            item["export"] = outputs[fmt.value]
        except Exception as exc:
            item["export_error"] = str(exc)
            item["doc_pages_len"] = len(doc.pages)
        summary["formats"][fmt.value] = item

    return summary


def main() -> None:
    image_path = ROOT / "tests" / "fixtures" / "test_en.png"
    with TemporaryDirectory(prefix="paddleocr_scan_smoke_") as tmp:
        out_dir = Path(tmp)
        (out_dir / "image").mkdir(parents=True, exist_ok=True)

        summary = {
            "image_scan": _run_image_flow(image_path, out_dir),
            "pdf_scan": _run_scanned_pdf_flow(image_path, out_dir),
        }
        print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
